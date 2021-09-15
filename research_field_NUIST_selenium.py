# ========#
#  设置  #
# ========#
# -*- coding:utf-8 -*-
import os, shutil
import time
import re
from openpyxl import load_workbook
from selenium import webdriver


url = 'https://esi.clarivate.com/'
chrome_path = os.getcwd() + "/Application/chrome.exe"
root, FILENAME=os.path.split(os.path.abspath(__file__))
# 用于暂存文件
temp_files = root + "\\temp_files"
if not os.path.exists(temp_files):
    os.mkdir(temp_files)

url = 'https://esi.clarivate.com/'
url2 = 'http://apps.webofknowledge.com/InboundService.do?product=WOS&Func=Frame&DestFail=http%3A%2F%2Fwww.webofknowledge.com%3FDestParams%3DUT%253DWOS%25253A000359216600008%2526customersID%253DInCites%2526smartRedirect%253Dyes%2526action%253Dretrieve%2526mode%253DFullRecord%2526product%253DCEL%26SrcAuth%3DInCites%26SrcApp%3DTSM_TEST%26DestApp%3DCEL%26e%3DvwtjxiLuDPrhqkxGzMeEeDiukRHn%252FfEx%252FpT5qofj3%252Boj8KOdkGQGQA%253D%253D&SrcApp=TSM_TEST&SrcAuth=InCites&SID=5FqAZgckxxK2sxI7pLR&customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&Init=Yes&action=retrieve&UT=WOS%3A000359216600008'
DE_url = ['https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=239&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=highlyCited&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls',
                 'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=15&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=Hot&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls']
#0:HCP,1:HP
TYPE = ['HCP','HP']
Headers={
         "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36",
         "Accept-Language": "zh-CN,zh;q=0.9"
         }
month_short = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
               'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
note_replace = {

    }
years = time.localtime(time.time()).tm_year


# ========#
#  函数   #
# ========#

# 打印信息时加上时间
def log_console(str):
    print('{} {}'.format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), str))

def chromeInit(savePath=os.getcwd(), flag=False):
    options=webdriver.ChromeOptions()
    options._binary_location= os.getcwd() + "/Application/chrome.exe"
    options.add_argument("--disable-gpu")
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-extensions')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    if flag:
        options.add_argument('--headless')

    if savePath != '':
        prefs = {
            "profile.default_content_settings.popups": 0,
            "download.default_directory": savePath,
            "profile.default_content_settings": {'images': 2}}
        options.add_experimental_option("prefs", prefs)
    chrome=webdriver.Chrome(options=options, executable_path=os.getcwd() + "/Application/chromedriver.exe")
    chrome.implicitly_wait(10)
    log_console("Chrome init complete...")
    return chrome

# 将数组逐行写入txt
def write_arr(arr, file):
    f = open(file, 'w')
    f.write('\n'.join(arr))
    f.close()

# 逐行读取txt生成返回数组
def load_arr(file):
    arr = []
    f = open(file, 'r')
    text = f.read()
    arr = text.split('\n')
    return arr
    f.close()

# 获取最新文件（时间排序取最后）
def sort_file(path):
    """排序文件"""
    dir_lists = os.listdir(path)
    dir_lists.sort(key=lambda fn: os.path.getmtime(path + '\\' + fn))
    return (dir_lists[-1])

# 移动文件，顺便改名
def movefile(src_file, dst_file):
    if not os.path.isfile(src_file):
        log_console("{} not exist!".format(src_file))
    else:
        dst_path,dst_fname=os.path.split(dst_file)     #分离文件名和路径
        if not os.path.exists(dst_path):   os.makedirs(dst_path)  #创建路径

        for fname in os.listdir(dst_path):  # 同名删除
            if fname == dst_fname: os.remove(dst_path + '\\' + fname)

        shutil.move(src_file, dst_file)         #移动文件
        # log_console("move {0} -> {1}".format(src_file, dst_file))

# 删除文件夹中的所有文件
def clean_fold(path):
    for i in os.listdir(path):              # os.listdir(path_data)#返回一个列表，里面是当前目录下面的所有东西的相对路径
        file = path + "\\" + i              # 当前文件夹的下面的所有东西的绝对路径
        if os.path.isfile(file) == True:    # os.path.isfile判断是否为文件,如果是文件,就删除.如果是文件夹.递归给del_file.
            os.remove(file)
        else:
            clean_fold(file)

# 用js点击
def js_click(driver, element):
    driver.execute_script("arguments[0].click();", element)

# 循环点击，直到下载完成
def loop_click(element):
    element.click()
    time.sleep(0.1)     # 等待下载完成
    if len(os.listdir(temp_files)) == 0:
        log_console('再点一次')
        loop_click(element)

def loop_js_click(driver, element):
    driver.execute_script("arguments[0].click();", element)
    time.sleep(1)       # 等待下载完成
    if len(os.listdir(temp_files)) == 0:
        log_console('再点一次')
        loop_js_click(driver, element)


def getUrlO(name):
    SID = '5FqAZgckxxK2sxI7pLR'
    url = 'http://apps.webofknowledge.com/InboundService.do?product=WOS&Func=Frame&DestFail=http%3A%2F%2Fwww.webofknowledge.com%3FDestParams%3DUT%253DWOS%25253A000359216600008%2526customersID%253DInCites%2526smartRedirect%253Dyes%2526action%253Dretrieve%2526mode%253DFullRecord%2526product%253DCEL%26SrcAuth%3DInCites%26SrcApp%3DTSM_TEST%26DestApp%3DCEL%26e%3DvwtjxiLuDPrhqkxGzMeEeDiukRHn%252FfEx%252FpT5qofj3%252Boj8KOdkGQGQA%253D%253D&SrcApp=TSM_TEST&SrcAuth=InCites&SID='+SID+'&customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&Init=Yes&action=retrieve&UT=WOS%3A'+name
    return url

def getUrl(name):
    url = 'http://gateway.webofknowledge.com/gateway/Gateway.cgi?GWVersion=2&SrcAuth=InCites&SrcApp=TSM_TEST&DestApp=WOS_CPL&DestLinkType=FullRecord&KeyUT='+name
    return url

def getUrl2O(name):
    SID = '5CgcjA1fRe1GUUrmdSV'
    url = 'http://cel.webofknowledge.com/InboundService.do?customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&product=CEL&Init=Yes&Func=Frame&action=retrieve&SrcApp=TSM_TEST&SrcAuth=InCites&SID='+SID+'&UT=WOS%3A'+name
    return url

def download(chrome, urls, dst_file):
    while True:
        try:
            log_console('访问此链接下载:{}'.format(urls))
            clean_fold(temp_files)
            chrome.get(urls)
        except Exception:
            log_console("网络错误（下载失败）\n")
            time.sleep(60)
            continue

        try:
            # 等待下载完成 确保中间文件(.tmp .crdownload)完全转好
            while (len(os.listdir(temp_files)) == 0
                   or os.listdir(temp_files)[0].split('.')[-1] == 'tmp'
                   or os.listdir(temp_files)[0].split('.')[-1] == 'crdownload'): time.sleep(0.1)

            movefile(temp_files + '\\' + sort_file(temp_files), dst_file)
            time.sleep(0.1)  # 稍微控制时间，防止反爬
            break
        except:
            log_console('文件移动转换错误... 重新下载...\n')

r'E:\BTH\ResearchFieldNUISTCrawler^(selenium^)\Application\chrome.exe --headless --disable-gpu --print-to-pdf=E:\BTH\ResearchFieldNUISTCrawler^(selenium^)\test.pdf E:\BTH\ResearchFieldNUISTCrawler^(selenium^)\2021.7\HCP\2-A SECURE AND DYNAMIC MULTI-KEYWORD RANKED SEARCH SCHEME OVER ENCRYPTED CLOUD DATA.html'

# 用chrome.exe将本地html转pdf
# src_file, dst_file 须为绝对路径
def html2pdf(chrome_path, src_file, dst_file):
    chrome_path = chrome_path.replace('(', '^(')\
        .replace(')', '^)')

    # 文件名里会有空格(cmd不能识别)，需要""
    src_file = r'"' + src_file + r'"'
    dst_file = r'"' + dst_file + r'"'

    cmd_str = chrome_path + ' ' + '--headless --disable-gpu --print-to-pdf=' + dst_file + ' ' + src_file
    cmd_str = cmd_str.replace('/', '\\')
    os.system(cmd_str)
    log_console(cmd_str)

first_enter = True
def save_pdf(chrome, urls, dst_path, title):
    while True:
        try:
            log_console('访问此链接下载html:{}'.format(urls))
            chrome.get(urls)

            # 如果是第一次打开 web of science 链接 网页些不必要的提示信息
            global first_enter
            if first_enter == True:
                time.sleep(0.5)
                chrome.get(urls)
                first_enter = False

            text = chrome.page_source
            break
        except Exception:
            log_console("网络错误（下载失败）\n")
            time.sleep(60)
            continue

    with open(dst_path + title + '.html', 'w', encoding='utf-8') as f:
        log_console(dst_path + title + '.html')
        f.write('<head><meta charset="UTF-8"></head>' + text)
    time.sleep(0.5)     # 防反爬

    html2pdf(chrome_path, dst_path + title + '.html', dst_path + title + '.pdf')

#=========#
# 主函数  #
#=========#
if __name__ == "__main__":
    log_console('已启动自动关机，将在21:55关机。')
    chrome = chromeInit(savePath=temp_files)
    log_console('正在加载...')
    while True:
        try:
            chrome.get(url)
            break
        except:
            log_console("网络错误（esi加载失败）")
            time.sleep(60)

    update_str = chrome.find_element_by_css_selector("#updateDateDatasetESI").text
    t = update_str.split('updated ')[1]
    t = t.split(' ')[0]
    month = month_short[t]
    log_console('当前更新至{}月'.format(month))

    # 主循环
    for i in range(2):
        try:
            f = open('./log/{}.{}/{}.log'.format(years, month, TYPE[i]), 'r')
        except:
            process = 0
            if not os.path.exists('./{}.{}/{}/'.format(years,month,TYPE[i])):
                os.makedirs('./{}.{}/{}/'.format(years,month,TYPE[i]))
            download(chrome, DE_url[i], './{1}.{2}/{0}/{0}-{1}.{2}.xlsx'.format(TYPE[i], years, month))
            log_console('{}总表获取完成'.format(TYPE[i]))
        else:
            process = int(f.read())
            log_console('{}已收集至{}'.format(TYPE[i],process))
            f.close()

        # 获取网页链接
        ws = load_workbook('./{1}.{2}/{0}/{0}-{1}.{2}.xlsx'.format(TYPE[i], years, month)).active
        total = ws.max_row - 8
        if process == total:
            log_console('{}已收集完成'.format(TYPE[i]))
            continue
        row_range = ws[7+process:len(ws['A'])-2]

        for item in row_range:
            title = item[3].value
            rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/ \ : * ? " < > |'
            title = re.sub(rstr, '', title)
            save_pdf(chrome, getUrlO(item[0].value.split(':')[1]), '{}/{}.{}/{}/'.format(os.getcwd(), years, month, TYPE[i])+'\\'+str(process + 1), '-' + title)
            if not os.path.exists('./log/{}.{}'.format(years, month)):
                os.makedirs('./log/{}.{}'.format(years, month))
            f = open('./log/{}.{}/{}.log'.format(years, month, TYPE[i]), 'w')
            process += 1
            f.write(str(process))
            f.close()
            log_console('{}进度{}/{}，{:.2f}%'.format(TYPE[i], process, total, process / total * 100))
            hour = time.localtime(time.time()).tm_hour
            minute = time.localtime(time.time()).tm_min
            if hour >= 21 and minute >= 55:
                os.system('shutdown -s')
        log_console('{}收集完成'.format(TYPE[i]))
    log_console('所有收集完成...\n')
    chrome.close()


    # log_console('进行pdf转换...')
    # data_root = '{}/{}.{}/'.format(os.getcwd(), years, month)
    # for path_1st in os.listdir(data_root):
    #     for file in os.listdir('{}{}'.format(data_root, path_1st)):
    #         if(file.split('.')[-1] == 'html'):
    #             html2pdf(chrome_path, '{}{}/{}'.format(data_root, path_1st, file), '{}{}/{}'.format(data_root, path_1st, file.replace('.html', '.pdf')))















